import os
import sys
import cv2
import json
import urllib.parse
import time
from datetime import datetime, timezone
from pathlib import Path

# External Libraries
from ollama import chat
from pymongo import MongoClient
import pymongo  # Added for database sorting
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font

# Google Drive API
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# MCP
from mcp.server.fastmcp import FastMCP

# ==========================================
# CONFIGURATION & INITIALIZATION
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(BASE_DIR)
load_dotenv(Path(__file__).parent / ".env")

MONGO_URI = os.getenv("MONGO_URI")
DB_NAME = "lifewood_db"
COLLECTION_NAME = "receipts"

# SCOPES = ["[https://www.googleapis.com/auth/drive.file](https://www.googleapis.com/auth/drive.file)"]
SCOPES = ["https://www.googleapis.com/auth/drive.file"]
CREDENTIALS_FILE = os.path.join(BASE_DIR, "credentials.json")
TOKEN_FILE = os.path.join(BASE_DIR, "token.json")
FOLDER_ID = os.getenv("GDRIVE_FOLDER_ID")

def log(msg: str):
    sys.stderr.write(f"[OcrReceipt] {msg}\n")
    sys.stderr.flush()

# ==========================================
# EXCEL BATCH MODULE
# ==========================================
def generate_merged_spreadsheet(
    all_items: list, sender: str, grand_total: float
) -> str:
    """Generates a single Excel file grouping items from multiple receipts."""
    log("Building consolidated Excel spreadsheet...")
    samples_dir = os.path.join(BASE_DIR, "samples")
    os.makedirs(samples_dir, exist_ok=True)

    file_path = os.path.join(samples_dir, f"receipts_merged_{sender}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Receipts"

    # Headers (Flat Layout matching Rose Pharmacy template)
    headers = ["Merchant Name", "Date", "Item Description", "Price"]
    ws.append(headers)

    for col in ["A1", "B1", "C1", "D1"]:
        ws[col].font = Font(bold=True)

    # Append all extracted items (repeats merchant and date per row)
    for item in all_items:
        ws.append([item["merchant"], item["date"], item["description"], item["price"]])

    # Add Grand Total Row directly below the items
    ws.append(["", "", "GRAND TOTAL:", grand_total])

    max_row = ws.max_row
    ws[f"C{max_row}"].font = Font(bold=True)
    ws[f"D{max_row}"].font = Font(bold=True)

    # Adjust column widths for clean formatting
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 15

    wb.save(file_path)
    return file_path

# ==========================================
# DATABASE MODULE
# ==========================================
def save_batch_to_mongo(batch_record: dict) -> str:
    """Saves a unified batch document containing all receipts to MongoDB."""
    if not MONGO_URI:
        raise ValueError("MONGO_URI not found in .env")

    client = MongoClient(MONGO_URI)
    db = client[DB_NAME]
    collection = db[COLLECTION_NAME]

    result = collection.insert_one(batch_record)
    client.close()
    return str(result.inserted_id)

# ==========================================
# GOOGLE DRIVE MODULE
# ==========================================
def get_google_auth():
    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, "w") as token:
            token.write(creds.to_json())
    return creds

def _upload(file_path: str, file_name: str, mimetype: str) -> str:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    if not FOLDER_ID:
        raise ValueError("GDRIVE_FOLDER_ID not found in .env")

    creds = get_google_auth()
    service = build("drive", "v3", credentials=creds)

    metadata = {"name": file_name, "parents": [FOLDER_ID]}
    media = MediaFileUpload(file_path, mimetype=mimetype, resumable=True)

    file = (
        service.files()
        .create(body=metadata, media_body=media, fields="id, webViewLink")
        .execute()
    )
    return file.get("webViewLink")

def upload_image_to_drive(image_path: str) -> str:
    name = f"Receipt_{os.path.basename(image_path)}"
    return _upload(image_path, name, "image/jpeg")

def upload_file_to_drive(
    file_path: str,
    mimetype: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
) -> str:
    name = os.path.basename(file_path)
    return _upload(file_path, name, mimetype)

# ==========================================
# VISION / OCR MODULE
# ==========================================
def preprocess_image(image_path: str) -> str:
    img = cv2.imread(image_path)
    if img is None:
        raise FileNotFoundError(f"Could not read image at: {image_path}")

    samples_dir = os.path.join(BASE_DIR, "samples")
    os.makedirs(samples_dir, exist_ok=True)

    original_filename = os.path.basename(image_path)
    output_path = os.path.join(samples_dir, f"cleaned_{original_filename}")

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    cleaned = cv2.adaptiveThreshold(
        gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2
    )
    cv2.imwrite(output_path, cleaned)
    return output_path

def compress_for_drive(image_path: str) -> str:
    img = cv2.imread(image_path)
    if img is None:
        return image_path

    max_width = 1200
    h, w = img.shape[:2]
    if w > max_width:
        ratio = max_width / w
        new_dim = (max_width, int(h * ratio))
        img = cv2.resize(img, new_dim, interpolation=cv2.INTER_AREA)

    samples_dir = os.path.join(BASE_DIR, "samples")
    os.makedirs(samples_dir, exist_ok=True)

    original_filename = os.path.basename(image_path)
    output_path = os.path.join(samples_dir, f"drive_compressed_{original_filename}")

    cv2.imwrite(output_path, img, [cv2.IMWRITE_JPEG_QUALITY, 70])
    return output_path

def extract_receipt_data(image_path: str) -> dict:
    """Sends the RAW image to the OCR model and safely parses the JSON."""
    system_prompt = """You are an expert receipt OCR engine. Extract data into EXACT JSON. 
    - Return ONLY valid JSON, no explanations, no markdown formatting, no backticks.
    - For missing/illegible fields use null.
    - Dates in YYYY-MM-DD. Time in HH:MM.
    - Prices as numbers (no currency symbols).
    - currency: "PHP", "USD", etc. or null.
    - items: list of {"description": str, "price": number}. Include ALL items.
    Schema:
    {
    "merchant_name": "string",
    "date": "string or null",
    "time": "string or null",
    "total_amount": number or null,
    "currency": "string or null",
    "items": [{"description": "string", "price": number}, ...]
    }
    """
    response = chat(
        model="qwen3-vl:8b-instruct",
        format="json",
        messages=[
            {
                "role": "user",
                "content": system_prompt,
                "images": [image_path],
            }
        ],
        options={
            "temperature": 0,
            "num_ctx": 16384,
            "keep_alive": "30m",
        },
    )

    raw_text = response.message.content.strip()

    if raw_text.startswith("```json"):
        raw_text = raw_text[7:]
    elif raw_text.startswith("```"):
        raw_text = raw_text[3:]

    if raw_text.endswith("```"):
        raw_text = raw_text[:-3]

    raw_text = raw_text.strip()

    try:
        receipt_data = json.loads(raw_text)
    except Exception as e:
        log(f"Failed to parse JSON. Raw model output was: {raw_text}")
        return {"error": "Failed to parse model output into structured JSON."}

    if not receipt_data.get("merchant_name") or not receipt_data.get("total_amount"):
        return {
            "error": "Image cannot be clearly read. Please scan again with better lighting."
        }

    return receipt_data

# ==========================================
# FAST MCP APP & TOOLS
# ==========================================
mcp = FastMCP("OcrReceipt")

def resolve_path(image_path: str) -> str:
    if os.path.exists(image_path):
        return image_path

    filename = os.path.basename(urllib.parse.urlparse(image_path).path)
    if not filename:
        filename = image_path

    home = os.path.expanduser("~")
    possible_paths = [
        os.path.join(home, ".openclaw", "media", "inbound", filename),
        os.path.join(home, ".openclaw", "media", filename),
        os.path.join("C:\\Users\\R3liq\\.openclaw", "media", "inbound", filename),
        os.path.join("C:\\Users\\R3liq\\.openclaw", "media", filename),
        os.path.join(BASE_DIR, filename),
    ]

    for path in possible_paths:
        if os.path.exists(path):
            return path
    return None

@mcp.tool()
def update_receipt_status(sender_name: str, status: str) -> str:
    """
    Updates the database status of the user's most recent receipt batch.
    Valid statuses: 'Confirmed', 'Needs Checking'
    """
    client = MongoClient(MONGO_URI)
    db = client[DB_NAME]
    collection = db[COLLECTION_NAME]

    latest = collection.find_one(
        {"sender_name": sender_name}, sort=[("createdAt", pymongo.DESCENDING)]
    )

    if not latest:
        client.close()
        return f"Could not find any recent receipts for {sender_name} to update."

    collection.update_one({"_id": latest["_id"]}, {"$set": {"status": status}})
    client.close()

    if status == "Confirmed":
        return f"✅ Database record for recent batch marked as *Confirmed*."
    else:
        return f"⚠️ Database record for recent batch flagged for *Checking*."

@mcp.tool()
def process_receipt(image_path: str, sender_name: str = "User") -> str:
    """
    Analyzes one OR MORE receipt images synchronously.
    CRITICAL LLM INSTRUCTION: Just pass the SINGLE most recent 'media://inbound/...' URI from the user's message into the `image_path` parameter. The Python script will automatically locate all other images the user sent in that batch.
    """
    paths = [p.strip() for p in image_path.split(",")]
    valid_paths = [resolve_path(p) for p in paths if resolve_path(p)]

    if not valid_paths:
        return "❌ Could not locate any image files. Please ensure they are downloaded."

    # ==========================================
    # AUTO-SCOOP LOGIC
    # Bypasses the LLM entirely and finds all images 
    # saved to the inbound folder in the last 3 minutes.
    # ==========================================
    final_paths = set(valid_paths)
    reference_file = valid_paths[0]

    try:
        inbound_dir = os.path.dirname(reference_file)
        ref_time = os.path.getmtime(reference_file)

        if os.path.exists(inbound_dir):
            for filename in os.listdir(inbound_dir):
                if filename.lower().endswith(('.png', '.jpg', '.jpeg')):
                    filepath = os.path.join(inbound_dir, filename)
                    file_time = os.path.getmtime(filepath)
                    # If file was modified within 180 seconds (3 mins) of our reference file, scoop it!
                    if abs(ref_time - file_time) <= 180:
                        final_paths.add(filepath)
    except Exception as e:
        log(f"Auto-scoop failed: {e}")

    valid_paths = list(final_paths)
    log(f"Auto-scooped a total of {len(valid_paths)} images for processing.")

    # ==========================================

    all_extracted_items = []
    total_sum = 0.0

    batch_mongo_record = {
        "sender_name": sender_name,
        "status": "Confirmed",
        "source": "WhatsApp OpenClaw",
        "createdAt": datetime.now(timezone.utc),
        "receipts": [],
        "grand_total": 0.0,
        "excel_link": None
    }

    final_reply = [f"✅ *Processed {len(valid_paths)} receipt(s) for {sender_name}!*\n"]

    for path in valid_paths:
        log(f"Extracting: {path}")
        receipt_data = extract_receipt_data(path)

        if "error" in receipt_data:
            final_reply.append(f"⚠️ Error reading an image: {receipt_data['error']}\n")
            continue

        drive_link = None
        try:
            colored_path = compress_for_drive(path)
            drive_link = upload_image_to_drive(colored_path)
        except Exception as e:
            log(f"Drive Image Upload Failed: {e}")

        merchant = receipt_data.get("merchant_name", "Unknown Merchant")
        date = receipt_data.get("date", "Unknown Date")
        currency = receipt_data.get("currency", "PHP")

        final_reply.append(f"🏪 *Merchant:* {merchant}")
        final_reply.append(f"📅 *Date:* {date}")
        if drive_link:
            final_reply.append(f"🖼️ *Image Backup:*\n{drive_link}")
        final_reply.append("") 

        receipt_entry = {
            "merchant_name": merchant,
            "date": date,
            "time": receipt_data.get("time"),
            "total_amount": receipt_data.get("total_amount"),
            "currency": currency,
            "drive_link": drive_link,
            "items": []
        }

        for item in receipt_data.get("items", []):
            try:
                clean_price = float(str(item.get("price", 0)).replace(",", "").strip())
            except:
                clean_price = 0.0

            all_extracted_items.append(
                {
                    "merchant": merchant,
                    "date": date,
                    "description": item.get("description", "Unknown Item"),
                    "price": clean_price,
                }
            )
            total_sum += clean_price
            receipt_entry["items"].append({"description": item.get("description", "Unknown Item"), "price": clean_price})

        batch_mongo_record["receipts"].append(receipt_entry)

    batch_mongo_record["grand_total"] = total_sum

    if all_extracted_items:
        try:
            excel_path = generate_merged_spreadsheet(all_extracted_items, sender_name, total_sum)
            excel_link = upload_file_to_drive(excel_path)
            batch_mongo_record["excel_link"] = excel_link

            final_reply.append(f"💰 *Grand Total:* PHP {total_sum}")
            final_reply.append(f"📊 *Excel Cloud Report:*\n{excel_link}\n")
        except Exception as e:
            log(f"Excel Error: {e}")
            final_reply.append(f"❌ Excel Error: {e}\n")

    try:
        save_batch_to_mongo(batch_mongo_record)
    except Exception as e:
        log(f"Mongo Error: {e}")

    # Aggressive Anti-Markdown Prompt
    final_reply.append(
        "\n---\n"
        "CRITICAL SYSTEM DIRECTIVE FOR LLM: \n"
        "You MUST output the text above exactly as is. \n"
        "DO NOT wrap your response in markdown backticks (```). \n"
        "DO NOT use code blocks or HTML tags. If you use a code block, you will crash the WhatsApp plugin. \n"
        "Reply with raw, plain text starting immediately with the ✅ emoji."
    )

    return "\n".join(final_reply)

# ==========================================
# MAIN EXECUTION
# ==========================================
if __name__ == "__main__":
    if len(sys.argv) > 1:
        test_images = sys.argv[1]
        print(f"--- Running Local CLI Batch Test ---")
        result = process_receipt(test_images, sender_name="Local_Tester")
        print("\n=== FINAL PIPELINE OUTPUT ===")
        print(result)
    else:
        mcp.run(transport="stdio")
