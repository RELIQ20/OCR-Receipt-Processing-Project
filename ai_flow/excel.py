import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def log(msg: str):
    sys.stderr.write(f"[OcrReceipt] {msg}\n")
    sys.stderr.flush()

def generate_merged_spreadsheet(all_items: list, sender: str, grand_total: float) -> str:
    """Generates a single Excel file grouping items from multiple receipts."""
    log("Building consolidated Excel spreadsheet...")
    samples_dir = os.path.join(BASE_DIR, "samples")
    os.makedirs(samples_dir, exist_ok=True)

    file_path = os.path.join(samples_dir, f"receipts_merged_{sender}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Receipts"

    # Headers (Flat Layout matching Rose Pharmacy template)
    headers = ["Merchant Name", "Date", "Item Description", "Quantity", "Price"]
    ws.append(headers)

    for col in ["A1", "B1", "C1", "D1", "E1"]:
        ws[col].font = Font(bold=True)

    # Append all extracted items (repeats merchant and date per row)
    for item in all_items:
        ws.append([item["merchant"], item["date"], item["description"], item.get("quantity", 1), item["price"]])

    # Add Grand Total Row directly below the items
    ws.append(["", "", "", "GRAND TOTAL:", grand_total])

    max_row = ws.max_row
    ws[f"D{max_row}"].font = Font(bold=True)
    ws[f"E{max_row}"].font = Font(bold=True)

    # Adjust column widths for clean formatting
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 15

    wb.save(file_path)
    return file_path
